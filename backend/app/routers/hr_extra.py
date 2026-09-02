"""الإشعارات، سجل التدقيق، وثائق الموظفين، والاستيراد الجماعي."""
from __future__ import annotations

import csv
import io
import secrets
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Response, UploadFile
from sqlalchemy import desc, select
from sqlalchemy.orm import Session

from ..config import MAX_UPLOAD_BYTES, UPLOAD_DIR
from ..database import get_db
from ..models import (
    AuditLog,
    Department,
    Employee,
    EmployeeDocument,
    EmployeeStatus,
    Notification,
    Role,
    Shift,
    User,
)
from ..schemas import AuditLogOut, DocumentIn, DocumentOut, ImportReport, NotificationOut
from ..security import can_view_employee, get_current_user, require_admin, require_hr
from ..services import audit, notifications, settings_store

router = APIRouter(prefix="/api", tags=["hr-extra"])

DOC_ATTACHMENTS = {".pdf", ".png", ".jpg", ".jpeg", ".webp"}
IMPORT_COLUMNS = ["رقم الموظف", "الاسم", "الإدارة", "المسمى الوظيفي", "الجوال", "البريد",
                  "الهوية", "تاريخ التعيين", "الراتب الأساسي", "الوردية"]


# ------------------------------ الإشعارات ------------------------------
@router.get("/notifications", response_model=list[NotificationOut])
def list_notifications(
    unread_only: bool = False,
    limit: int = Query(50, le=200),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    stmt = select(Notification).where(Notification.user_id == user.id)
    if unread_only:
        stmt = stmt.where(Notification.is_read.is_(False))
    return db.scalars(stmt.order_by(desc(Notification.created_at), desc(Notification.id)).limit(limit)).all()


@router.get("/notifications/unread-count")
def unread_count(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    rows = db.scalars(
        select(Notification).where(Notification.user_id == user.id, Notification.is_read.is_(False))
    ).all()
    return {"count": len(rows)}


@router.post("/notifications/{notification_id}/read", response_model=NotificationOut)
def mark_read(notification_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    row = db.get(Notification, notification_id)
    if not row or row.user_id != user.id:
        raise HTTPException(status_code=404, detail="الإشعار غير موجود")
    row.is_read = True
    db.commit()
    db.refresh(row)
    return row


@router.post("/notifications/read-all")
def mark_all_read(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    rows = db.scalars(
        select(Notification).where(Notification.user_id == user.id, Notification.is_read.is_(False))
    ).all()
    for row in rows:
        row.is_read = True
    db.commit()
    return {"ok": True, "count": len(rows)}


# ------------------------------ سجل التدقيق ------------------------------
@router.get("/audit-logs", response_model=list[AuditLogOut])
def list_audit(
    entity: str | None = None,
    action: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    limit: int = Query(200, le=1000),
    db: Session = Depends(get_db),
    _: User = Depends(require_hr),
):
    stmt = select(AuditLog)
    if entity:
        stmt = stmt.where(AuditLog.entity == entity)
    if action:
        stmt = stmt.where(AuditLog.action == action)
    if date_from:
        stmt = stmt.where(AuditLog.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        stmt = stmt.where(AuditLog.created_at <= datetime.combine(date_to, datetime.max.time()))
    rows = db.scalars(stmt.order_by(desc(AuditLog.created_at), desc(AuditLog.id)).limit(limit)).all()
    users = {u.id: u.username for u in db.scalars(select(User)).all()}
    return [
        AuditLogOut(
            id=r.id,
            user_id=r.user_id,
            username=users.get(r.user_id or 0),
            action=r.action,
            action_label=audit.ACTION_LABELS.get(r.action, r.action),
            entity=r.entity,
            entity_label=audit.ENTITY_LABELS.get(r.entity, r.entity),
            entity_id=r.entity_id,
            detail=r.detail,
            created_at=r.created_at,
        )
        for r in rows
    ]


# ------------------------------ وثائق الموظفين ------------------------------
def document_out(doc: EmployeeDocument) -> DocumentOut:
    days_left = (doc.expiry_date - date.today()).days if doc.expiry_date else None
    return DocumentOut(
        id=doc.id,
        employee_id=doc.employee_id,
        employee_name=doc.employee.full_name if doc.employee else None,
        employee_code=doc.employee.code if doc.employee else None,
        doc_type=doc.doc_type,
        number=doc.number,
        issue_date=doc.issue_date,
        expiry_date=doc.expiry_date,
        file_path=doc.file_path,
        note=doc.note,
        days_left=days_left,
    )


@router.get("/documents", response_model=list[DocumentOut])
def list_documents(
    employee_id: int | None = None,
    expiring_days: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    stmt = select(EmployeeDocument)
    if user.role == Role.employee:
        stmt = stmt.where(EmployeeDocument.employee_id == (user.employee_id or 0))
    if employee_id:
        if not can_view_employee(user, employee_id, db):
            raise HTTPException(status_code=403, detail="لا تملك صلاحية عرض هذا الموظف")
        stmt = stmt.where(EmployeeDocument.employee_id == employee_id)
    rows = db.scalars(stmt.order_by(EmployeeDocument.expiry_date)).all()
    result = [document_out(d) for d in rows]
    if expiring_days is not None:
        result = [d for d in result if d.days_left is not None and d.days_left <= expiring_days]
    return result


@router.post("/documents", response_model=DocumentOut, status_code=201)
def create_document(payload: DocumentIn, db: Session = Depends(get_db), user: User = Depends(require_hr)):
    if not db.get(Employee, payload.employee_id):
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    doc = EmployeeDocument(**payload.model_dump())
    db.add(doc)
    db.flush()
    audit.log(db, user, "create", "document", doc.id, f"{doc.doc_type}", commit=False)
    db.commit()
    db.refresh(doc)
    return document_out(doc)


@router.post("/documents/{document_id}/file", response_model=DocumentOut)
def upload_document_file(
    document_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_hr),
):
    doc = db.get(EmployeeDocument, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="الوثيقة غير موجودة")
    suffix = "." + (file.filename or "").rsplit(".", 1)[-1].lower() if "." in (file.filename or "") else ""
    if suffix not in DOC_ATTACHMENTS:
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم (PDF أو صورة فقط)")
    content = file.file.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail="حجم الملف أكبر من الحد المسموح")
    name = f"doc_{doc.id}_{secrets.token_hex(6)}{suffix}"
    (UPLOAD_DIR / name).write_bytes(content)
    doc.file_path = name
    db.commit()
    db.refresh(doc)
    return document_out(doc)


@router.delete("/documents/{document_id}")
def delete_document(document_id: int, db: Session = Depends(get_db), user: User = Depends(require_hr)):
    doc = db.get(EmployeeDocument, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="الوثيقة غير موجودة")
    db.delete(doc)
    audit.log(db, user, "delete", "document", document_id, commit=False)
    db.commit()
    return {"ok": True}


@router.post("/documents/scan-expiring")
def scan_expiring(db: Session = Depends(get_db), user: User = Depends(require_hr)):
    """يفحص الوثائق المقاربة على الانتهاء ويرسل إشعارات للموارد البشرية والموظفين."""
    days = settings_store.get_int(db, "document_alert_days", 30)
    today = date.today()
    sent = 0
    for doc in db.scalars(select(EmployeeDocument).where(EmployeeDocument.expiry_date.is_not(None))).all():
        left = (doc.expiry_date - today).days
        if left > days:
            continue
        name = doc.employee.full_name if doc.employee else ""
        title = (
            f"انتهت صلاحية {doc.doc_type} للموظف {name}"
            if left < 0
            else f"{doc.doc_type} للموظف {name} تنتهي خلال {left} يوم"
        )
        notifications.notify_roles(
            db, [Role.admin, Role.hr], title,
            body=f"تاريخ الانتهاء: {doc.expiry_date}", category="document",
            link_page="documents", commit=False,
        )
        notifications.notify_employee(
            db, doc.employee_id, title, body=f"تاريخ الانتهاء: {doc.expiry_date}",
            category="document", link_page="documents", commit=False,
        )
        sent += 1
    db.commit()
    return {"ok": True, "documents": sent, "message": f"تم إرسال تنبيهات لـ {sent} وثيقة"}


# ------------------------------ استيراد الموظفين ------------------------------
@router.get("/employees-import-template.csv")
def import_template(_: User = Depends(require_hr)):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(IMPORT_COLUMNS)
    writer.writerow(["1001", "عبدالله محمد", "تقنية المعلومات", "مطور", "0500000000",
                     "a@example.com", "1012345678", "2024-01-15", "9000", "الوردية الصباحية"])
    return Response(
        "﻿" + buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=employees_template.csv"},
    )


def _rows_from_upload(filename: str, content: bytes) -> list[list[str]]:
    """يقرأ صفوف ملف Excel أو CSV."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        return [
            ["" if cell is None else str(cell).strip() for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    text = content.decode("utf-8-sig", errors="ignore")
    delimiter = ";" if text.count(";") > text.count(",") else ","
    return [[c.strip() for c in row] for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _parse_date(value: str):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            continue
    return None


@router.post("/employees/import", response_model=ImportReport)
def import_employees(
    file: UploadFile = File(...),
    update_existing: bool = Form(True),
    db: Session = Depends(get_db),
    user: User = Depends(require_hr),
):
    """استيراد جماعي للموظفين من ملف Excel أو CSV بالأعمدة الموضحة في القالب."""
    content = file.file.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail="حجم الملف أكبر من الحد المسموح")
    rows = _rows_from_upload(file.filename or "", content)
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="الملف فارغ أو لا يحتوي صفوف بيانات")

    departments = {d.name: d for d in db.scalars(select(Department)).all()}
    shifts = {s.name: s for s in db.scalars(select(Shift)).all()}
    existing = {e.code: e for e in db.scalars(select(Employee)).all()}
    report = ImportReport()

    for index, row in enumerate(rows[1:], start=2):
        if not any((c or "").strip() for c in row):
            continue
        cells = list(row) + [""] * (10 - len(row))
        code, name = (cells[0] or "").strip(), (cells[1] or "").strip()
        if not code or not name:
            report.errors.append(f"السطر {index}: رقم الموظف أو الاسم مفقود")
            report.skipped += 1
            continue

        dep_name = (cells[2] or "").strip()
        department = departments.get(dep_name)
        if dep_name and department is None:
            department = Department(name=dep_name)
            db.add(department)
            db.flush()
            departments[dep_name] = department
        shift = shifts.get((cells[9] or "").strip())

        try:
            salary = float((cells[8] or "0").replace(",", "") or 0)
        except ValueError:
            salary = 0.0

        data = dict(
            full_name=name,
            department_id=department.id if department else None,
            job_title=(cells[3] or "").strip() or None,
            phone=(cells[4] or "").strip() or None,
            email=(cells[5] or "").strip() or None,
            national_id=(cells[6] or "").strip() or None,
            hire_date=_parse_date(cells[7] or ""),
            basic_salary=salary,
            shift_id=shift.id if shift else None,
        )

        employee = existing.get(code)
        if employee is None:
            employee = Employee(code=code, **data)
            db.add(employee)
            db.flush()
            existing[code] = employee
            report.created += 1
        elif update_existing:
            for key, value in data.items():
                if value not in (None, "", 0.0) or key in ("basic_salary",):
                    setattr(employee, key, value)
            report.updated += 1
        else:
            report.skipped += 1

    audit.log(
        db, user, "import", "employee", None,
        f"أُنشئ {report.created} وحُدّث {report.updated}", commit=False,
    )
    db.commit()
    report.message = (
        f"تم إنشاء {report.created} موظف وتحديث {report.updated}"
        + (f" وتخطي {report.skipped}" if report.skipped else "")
    )
    return report
